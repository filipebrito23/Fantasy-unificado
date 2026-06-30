from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from data_loader import load_workbook_data, SEASONS
from transforms import (
    SEASON_LABELS,
    build_picks_view,
    format_picks_for_display,
    get_team_options,
    get_visible_seasons,
    build_roster_view,
    format_roster_for_display,
    calculate_main_totals,
    calculate_dev_totals,
    summarize_positions,
    summarize_picks_by_year,
)

DEFAULT_FILE = Path("roster.xlsx")
TX_SHEET = "transactions"
TX_ITEMS_SHEET = "transactionitems"


def require_login_v5():
    if "user_v5" not in st.session_state or not st.session_state.user_v5:
        st.warning("Faça login para acessar esta página.")
        st.stop()
    return st.session_state.user_v5


@st.cache_data
def cached_load(file_path: str):
    return load_workbook_data(file_path)


def currency(v) -> str:
    if v is None or (isinstance(v, str) and not v.strip()):
        return "-"

    n = pd.to_numeric(pd.Series([v]), errors="coerce").iloc[0]

    if pd.isna(n):
        return str(v)

    return f"US$ {float(n):,.2f}"


def format_salary_columns(df: pd.DataFrame, visible_seasons: list[str]) -> pd.DataFrame:
    if df.empty:
        return df

    out = df.copy()
    for season in visible_seasons:
        label = SEASON_LABELS[season]
        if label in out.columns:
            out[label] = out[label].apply(currency)
    return out


def format_money_columns(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    if df.empty:
        return df

    out = df.copy()
    for col in cols:
        if col in out.columns:
            out[col] = out[col].apply(currency)
    return out


def display_table(df: pd.DataFrame):
    st.dataframe(df, use_container_width=True, hide_index=True)


def get_next_id(df: pd.DataFrame, col: str, start: int = 1) -> int:
    if df.empty or col not in df.columns:
        return start
    s = pd.to_numeric(df[col], errors="coerce").dropna()
    return int(s.max()) + 1 if not s.empty else start


def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    out = out.where(pd.notna(out), None)
    return out


def load_sheet_df(wb, sheet_name: str) -> pd.DataFrame:
    real_name = find_sheet_name(wb, sheet_name)
    if real_name is None:
        return pd.DataFrame()
    ws = wb[real_name]
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows[1:], columns=rows[0])


def save_sheet_df(wb, sheet_name: str, df: pd.DataFrame):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)

    ws = wb.create_sheet(sheet_name)

    if df.empty:
        return

    df2 = normalize_df(df)
    df2 = df2.astype(object).where(pd.notna(df2), None)

    ws.append(list(df2.columns))
    for _, row in df2.iterrows():
        ws.append([None if pd.isna(v) else v for v in row.tolist()])


def roster_domain_ids(data, team_id: int) -> set:
    ids = set()
    for sheet in ["roster", "development"]:
        df = data.get(sheet, pd.DataFrame())
        if not df.empty and {"team_id", "player_id"}.issubset(df.columns):
            ids |= set(
                df.loc[df["team_id"].astype(int) == team_id, "player_id"]
                .dropna()
                .astype(int)
                .tolist()
            )
    return ids


def pick_domain_ids(data, team_id: int) -> set:
    picks = data.get("picks", pd.DataFrame())
    if picks.empty:
        return set()

    owner_cols = [c for c in picks.columns if "owner" in c.lower() or "team" in c.lower()]
    if not owner_cols:
        return set()

    owner_col = owner_cols[0]
    id_col = picks.columns[0]

    return set(
        picks.loc[picks[owner_col].astype(int) == team_id, id_col]
        .astype(str)
        .tolist()
    )


def validate_items(data, from_team_id: int, item_rows: list[dict]) -> list[str]:
    errors = []
    player_ids = roster_domain_ids(data, from_team_id)
    pick_ids = pick_domain_ids(data, from_team_id)

    for i, item in enumerate(item_rows, start=1):
        if item["item_type"] == "player" and item["asset_id"] not in player_ids:
            errors.append(f"Item {i}: jogador fora do domínio do time origem.")
        if item["item_type"] == "pick" and str(item["asset_id"]) not in pick_ids:
            errors.append(f"Item {i}: pick fora do domínio do time origem.")

    return errors


def append_transaction(file_path: str, tx_row: dict, item_rows: list[dict]):
    wb = load_workbook(file_path)

    tx_df = load_sheet_df(wb, TX_SHEET)
    items_df = load_sheet_df(wb, TX_ITEMS_SHEET)

    if tx_df.empty and TX_SHEET not in wb.sheetnames:
        tx_df = pd.DataFrame(columns=list(tx_row.keys()))

    if items_df.empty and TX_ITEMS_SHEET not in wb.sheetnames and item_rows:
        items_df = pd.DataFrame(columns=list(item_rows[0].keys()))

    tx_df = pd.concat([tx_df, pd.DataFrame([tx_row])], ignore_index=True)

    if item_rows:
        items_df = pd.concat([items_df, pd.DataFrame(item_rows)], ignore_index=True)

    save_sheet_df(wb, TX_SHEET, tx_df)
    if item_rows:
        save_sheet_df(wb, TX_ITEMS_SHEET, items_df)

    wb.save(file_path)


def update_rosters(file_path: str, tx_row: dict, item_rows: list[dict]):
    wb = load_workbook(file_path)

    roster_df = load_sheet_df(wb, "roster")
    dev_df = load_sheet_df(wb, "development")
    picks_df = load_sheet_df(wb, "picks")

    from_team_id = int(tx_row["from_team_id"])
    to_team_id = int(tx_row["to_team_id"])

    for item in item_rows:
        if item["item_type"] == "player":
            pid = int(item["asset_id"])

            if item["from_roster_type"] == "MAIN" and not roster_df.empty:
                mask = (
                    (roster_df["team_id"].astype(int) == from_team_id)
                    & (roster_df["player_id"].astype(int) == pid)
                )
                roster_df.loc[mask, "team_id"] = to_team_id

            elif item["from_roster_type"] == "DEV" and not dev_df.empty:
                mask = (
                    (dev_df["team_id"].astype(int) == from_team_id)
                    & (dev_df["player_id"].astype(int) == pid)
                )
                dev_df.loc[mask, "team_id"] = to_team_id

        elif item["item_type"] == "pick" and not picks_df.empty:
            id_col = picks_df.columns[0]
            owner_cols = [c for c in picks_df.columns if "owner" in c.lower() or "team" in c.lower()]
            if owner_cols:
                owner_col = owner_cols[0]
                mask = (
                    picks_df[id_col].astype(str).eq(str(item["asset_id"]))
                    & picks_df[owner_col].astype(int).eq(from_team_id)
                )
                picks_df.loc[mask, owner_col] = to_team_id

    save_sheet_df(wb, "roster", roster_df)
    save_sheet_df(wb, "development", dev_df)
    save_sheet_df(wb, "picks", picks_df)

    wb.save(file_path)


def build_red_flags(df: pd.DataFrame, visible_seasons: list[str]) -> pd.DataFrame:
    if df.empty:
        return df

    out = df.copy()

    for season in visible_seasons:
        sal_col = SEASON_LABELS[season]
        opt_col = f"TO_{season}"

        if sal_col not in out.columns or opt_col not in out.columns:
            continue

        mask = out[opt_col].astype(str).str.strip().str.lower().eq("sim")
        out[sal_col] = out[sal_col].apply(currency)
        out.loc[mask, sal_col] = out.loc[mask, sal_col].astype(str) + " 🔴"

    to_cols = [f"TO_{season}" for season in visible_seasons if f"TO_{season}" in out.columns]
    out = out.drop(columns=to_cols, errors="ignore")

    return out

def compact_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [
        str(c).strip().lower().replace("_", "").replace(" ", "")
        for c in out.columns
    ]
    return out

def build_transactions_history(
    tx_df: pd.DataFrame,
    items_df: pd.DataFrame,
    selected_team_id: int,
    team_lookup: dict,
    player_lookup: dict,
) -> pd.DataFrame:
    if tx_df is None or tx_df.empty:
        return pd.DataFrame()

    tx = compact_columns(tx_df)

    items = items_df.copy() if items_df is not None else pd.DataFrame()
    if not items.empty:
        items = compact_columns(items)

    if not {"fromteamid", "toteamid"}.issubset(tx.columns):
        return pd.DataFrame()

    team_mask = (
        tx["fromteamid"].astype(str).eq(str(selected_team_id))
        | tx["toteamid"].astype(str).eq(str(selected_team_id))
    )
    tx = tx.loc[team_mask].copy()

    if tx.empty:
        return pd.DataFrame()

    tx["from_team"] = tx["fromteamid"].map(team_lookup).fillna(tx["fromteamid"])
    tx["to_team"] = tx["toteamid"].map(team_lookup).fillna(tx["toteamid"])

    if items.empty or "transactionid" not in items.columns:
        tx["items_summary"] = "-"
    else:
        items = items.loc[
            items["transactionid"].astype(str).isin(tx["transactionid"].astype(str))
        ].copy()

        def describe_item(row):
            item_type = str(row.get("itemtype", "")).upper()
            asset_id = row.get("assetid")

            if item_type == "PLAYER":
                try:
                    pid = int(asset_id)
                except Exception:
                    pid = asset_id
                asset_label = player_lookup.get(pid, asset_id)
            else:
                asset_label = asset_id

            from_rt = row.get("fromrostertype")
            to_rt = row.get("torostertype")

            roster_part = ""
            if pd.notna(from_rt) or pd.notna(to_rt):
                roster_part = f" ({from_rt or '-'} → {to_rt or '-'})"

            return f"{item_type}: {asset_label}{roster_part}"

        if not items.empty:
            items["item_desc"] = items.apply(describe_item, axis=1)
            items_grouped = (
                items.groupby("transactionid")["item_desc"]
                .apply(lambda s: " | ".join(s.astype(str)))
                .reset_index()
            )
            tx = tx.merge(items_grouped, on="transactionid", how="left")
            tx = tx.rename(columns={"item_desc": "items_summary"})
        else:
            tx["items_summary"] = "-"

    tx = tx.rename(columns={
        "transactionid": "transaction_id",
        "transactiontype": "transaction_type",
        "transactiondate": "transaction_date",
        "initiatedby": "initiated_by",
    })

    preferred_cols = [
        "transaction_id",
        "transaction_date",
        "transaction_type",
        "season",
        "from_team",
        "to_team",
        "initiated_by",
        "status",
        "items_summary",
        "notes",
    ]
    existing_cols = [c for c in preferred_cols if c in tx.columns]

    sort_cols = [c for c in ["transaction_date", "transaction_id"] if c in tx.columns]
    if sort_cols:
        tx = tx.sort_values(by=sort_cols, ascending=False)

    return tx[existing_cols]

def format_totals(df: pd.DataFrame, money_cols: list[str]) -> pd.DataFrame:
    if df.empty:
        return df

    out = df.copy()
    for col in money_cols:
        if col in out.columns:
            out[col] = out[col].apply(currency)
    return out

def get_roster_column_order(df: pd.DataFrame, visible_seasons: list[str]) -> list[str]:
    ordered = ["Ordem", "Jogador", "Posição"]
    ordered += [
        SEASON_LABELS[season]
        for season in visible_seasons
        if SEASON_LABELS[season] in df.columns
    ]
    return [col for col in ordered if col in df.columns]


def get_roster_column_config(visible_seasons: list[str]) -> dict:
    config = {
        "Ordem": st.column_config.NumberColumn("Ordem", width="small", format="%d"),
        "Jogador": st.column_config.TextColumn("Jogador", width="large"),
        "Posição": st.column_config.TextColumn("Posição", width="small"),
    }

    for season in visible_seasons:
        label = SEASON_LABELS[season]
        config[label] = st.column_config.TextColumn(label, width="medium")

    return config


def get_picks_column_order(df: pd.DataFrame) -> list[str]:
    preferred = ["Pick", "Ano", "Round", "Time original", "Time atual"]
    return [col for col in preferred if col in df.columns]


def get_picks_column_config() -> dict:
    return {
        "Pick": st.column_config.TextColumn("Pick", width="medium"),
        "Ano": st.column_config.NumberColumn("Ano", width="small", format="%d"),
        "Round": st.column_config.NumberColumn("Round", width="small", format="%d"),
        "Time original": st.column_config.TextColumn("Time original", width="medium"),
        "Time atual": st.column_config.TextColumn("Time atual", width="medium"),
    }

user = require_login_v5()

st.title("Elencos Fantasy NBA")
st.caption("Elencos 2026-27")

if not DEFAULT_FILE.exists():
    st.error("Arquivo roster.xlsx não encontrado na pasta do projeto.")
    st.stop()

data = cached_load(str(DEFAULT_FILE))
teams = get_team_options(data["teams"])

if teams.empty:
    st.error("Nenhum time encontrado nos dados carregados.")
    st.stop()

if "teams_selected_team_name_v1" not in st.session_state:
    st.session_state["teams_selected_team_name_v1"] = teams["team_name"].tolist()[0]

if "teams_selected_start_season_v1" not in st.session_state:
    st.session_state["teams_selected_start_season_v1"] = SEASONS[0]

team_map = (
    data["teams"][["team_id", "team_name"]].drop_duplicates()
    if not data["teams"].empty
    else pd.DataFrame(columns=["team_id", "team_name"])
)
team_lookup = dict(zip(team_map["team_id"], team_map["team_name"])) if not team_map.empty else {}

player_lookup = (
    dict(zip(data["players"]["player_id"], data["players"]["player_name"]))
    if not data["players"].empty
    else {}
)

picks_df = data.get("picks", pd.DataFrame())
pick_id_col = picks_df.columns[0] if not picks_df.empty else None
pick_choices = picks_df[pick_id_col].astype(str).tolist() if pick_id_col else []

c1, c2 = st.columns([2, 1])

with c1:
    selected_team_name = st.selectbox(
        "Selecione o time",
        teams["team_name"].tolist(),
        key="teams_selected_team_name_v1",
    )

with c2:
    selected_start_season = st.selectbox(
        "Temporada inicial",
        SEASONS,
        format_func=lambda x: SEASON_LABELS[x],
        key="teams_selected_start_season_v1",
    )

selected_team_id = int(
    teams.loc[teams["team_name"] == selected_team_name, "team_id"].iloc[0]
)
visible_seasons = get_visible_seasons(selected_start_season)

main_team_df = data["roster"].loc[data["roster"]["team_id"] == selected_team_id].copy()
dev_team_df = data["development"].loc[data["development"]["team_id"] == selected_team_id].copy()

main_roster_raw = build_roster_view(
    main_team_df, data["players"], selected_team_id, "MAIN", visible_seasons
)
main_roster = format_roster_for_display(main_roster_raw, visible_seasons)
main_totals = calculate_main_totals(main_team_df, data["fines"], selected_team_id, visible_seasons)

dev_roster_raw = build_roster_view(
    dev_team_df, data["players"], selected_team_id, "DEV", visible_seasons
)
dev_roster = format_roster_for_display(dev_roster_raw, visible_seasons)
dev_totals = calculate_dev_totals(dev_team_df, visible_seasons)

picks_team_df = build_picks_view(data["picks"], data["teams"], selected_team_id)
picks_display = format_picks_for_display(picks_team_df)

main_position_counts = summarize_positions(main_roster)
dev_position_counts = summarize_positions(dev_roster)

main_positions_text = " | ".join([f"{pos}: {qty}" for pos, qty in main_position_counts.items()]) if main_position_counts else "-"
dev_positions_text = " | ".join([f"{pos}: {qty}" for pos, qty in dev_position_counts.items()]) if dev_position_counts else "-"

team_picks_df = picks_df.loc[picks_df["current_team_owner_id"] == selected_team_id].copy() if not picks_df.empty else pd.DataFrame()
pick_year_counts = summarize_picks_by_year(team_picks_df)
picks_display = team_picks_df.copy()

transactions_df = data.get(TX_SHEET, pd.DataFrame())
transaction_items_df = data.get(TX_ITEMS_SHEET, pd.DataFrame())

team_transactions_df = build_transactions_history(
    transactions_df,
    transaction_items_df,
    selected_team_id,
    team_lookup,
    player_lookup,
)

if not picks_display.empty:
    if "original_team_pick_id" in picks_display.columns:
        picks_display["Time original"] = picks_display["original_team_pick_id"].map(team_lookup).fillna(
            picks_display["original_team_pick_id"]
        )

    if "current_team_owner_id" in picks_display.columns:
        picks_display["Time atual"] = picks_display["current_team_owner_id"].map(team_lookup).fillna(
            picks_display["current_team_owner_id"]
        )

    rename_map = {
        "pick_id": "Pick",
        "year": "Ano",
        "round": "Round",
    }
    picks_display = picks_display.rename(columns=rename_map)


st.subheader(f"Elencos - {selected_team_name}")

tab_main, tab_dev, tab_picks, tab_transactions = st.tabs(["Principal", "Development", "Picks", "Transactions"])

with tab_main:
    c1, c2, c3, c4 = st.columns(4)
    if not main_totals.empty:
        current_main = main_totals.iloc[0].to_dict()
        c1.metric("Jogadores", len(main_roster))
        c2.metric("Salários", currency(current_main.get("Salários", 0.0)))
        c3.metric("Multas", currency(current_main.get("Multas", 0.0)))
        c4.metric("Cap restante", currency(current_main.get("Cap restante", 0.0)))

    st.caption(f"Posições: {main_positions_text}")

    display_main = build_red_flags(main_roster, visible_seasons)

    st.dataframe(
        display_main,
        use_container_width=True,
        hide_index=True,
        column_order=get_roster_column_order(display_main, visible_seasons),
        column_config=get_roster_column_config(visible_seasons),
    )

    with st.expander("Totalizadores do elenco principal", expanded=False):
        main_totals_display = main_totals.copy()
        for col in ["Salários", "Multas", "Cap restante"]:
            if col in main_totals_display.columns:
                main_totals_display[col] = main_totals_display[col].apply(currency)
        st.dataframe(main_totals_display, use_container_width=True, hide_index=True)

with tab_dev:
    c1, c2, c3 = st.columns(3)
    if not dev_totals.empty:
        current_dev = dev_totals.iloc[0].to_dict()
        c1.metric("Jogadores", len(dev_roster))
        c2.metric("Salários", currency(current_dev.get("Salários", 0.0)))
        c3.metric("Cap restante", currency(current_dev.get("Cap restante", 0.0)))

    st.caption(f"Posições: {dev_positions_text}")

    display_dev = build_red_flags(dev_roster, visible_seasons)

    st.dataframe(
        display_dev,
        use_container_width=True,
        hide_index=True,
        column_order=get_roster_column_order(display_dev, visible_seasons),
        column_config=get_roster_column_config(visible_seasons),
    )

    with st.expander("Totalizadores da development", expanded=False):
        dev_totals_display = dev_totals.copy()
        for col in ["Salários", "Cap restante"]:
            if col in dev_totals_display.columns:
                dev_totals_display[col] = dev_totals_display[col].apply(currency)
        st.dataframe(dev_totals_display, use_container_width=True, hide_index=True)

with tab_picks:
    total_picks = len(team_picks_df)

    if team_picks_df.empty:
        st.info("Esse time não possui picks cadastradas.")
    else:
        metric_cols = st.columns(len(pick_year_counts) + 1 if pick_year_counts else 1)
        metric_cols[0].metric("Total de picks", total_picks)

        for idx, (year, qty) in enumerate(sorted(pick_year_counts.items()), start=1):
            metric_cols[idx].metric(str(year), qty)

        st.dataframe(
            picks_display,
            use_container_width=True,
            hide_index=True,
            column_order=get_picks_column_order(picks_display),
            column_config=get_picks_column_config(),
        )

with tab_transactions:
    st.subheader("Histórico de transactions")

    if team_transactions_df.empty:
        st.info("Esse time ainda não possui transactions registradas.")
    else:
        filter_col1, filter_col2, filter_col3 = st.columns(3)

        type_options = ["Todas"]
        if "transaction_type" in team_transactions_df.columns:
            type_options += sorted(
                team_transactions_df["transaction_type"]
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )

        status_options = ["Todos"]
        if "status" in team_transactions_df.columns:
            status_options += sorted(
                team_transactions_df["status"]
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )

        season_options = ["Todas"]
        if "season" in team_transactions_df.columns:
            season_options += sorted(
                team_transactions_df["season"]
                .dropna()
                .astype(str)
                .unique()
                .tolist()
            )

        with filter_col1:
            selected_tx_type = st.selectbox(
                "Tipo",
                type_options,
                key="tx_history_type_v3",
            )

        with filter_col2:
            selected_tx_status = st.selectbox(
                "Status",
                status_options,
                key="tx_history_status_v3",
            )

        with filter_col3:
            selected_tx_season = st.selectbox(
                "Season",
                season_options,
                key="tx_history_season_v3",
            )

        filtered_tx = team_transactions_df.copy()

        if selected_tx_type != "Todas" and "transaction_type" in filtered_tx.columns:
            filtered_tx = filtered_tx.loc[
                filtered_tx["transaction_type"].astype(str) == selected_tx_type
            ]

        if selected_tx_status != "Todos" and "status" in filtered_tx.columns:
            filtered_tx = filtered_tx.loc[
                filtered_tx["status"].astype(str) == selected_tx_status
            ]

        if selected_tx_season != "Todas" and "season" in filtered_tx.columns:
            filtered_tx = filtered_tx.loc[
                filtered_tx["season"].astype(str) == selected_tx_season
            ]

        metric_cols = st.columns(3)
        metric_cols[0].metric("Transactions", len(filtered_tx))
        metric_cols[1].metric(
            "Tipos distintos",
            filtered_tx["transaction_type"].nunique()
            if "transaction_type" in filtered_tx.columns
            else 0,
        )
        metric_cols[2].metric(
            "Seasons",
            filtered_tx["season"].nunique()
            if "season" in filtered_tx.columns
            else 0,
        )

        display_cols = [
            col
            for col in [
                "transaction_id",
                "transaction_date",
                "transaction_type",
                "season",
                "from_team",
                "to_team",
                "initiated_by",
                "status",
                "items_summary",
                "notes",
            ]
            if col in filtered_tx.columns
        ]

        st.dataframe(
            filtered_tx[display_cols] if display_cols else filtered_tx,
            use_container_width=True,
            hide_index=True,
        )