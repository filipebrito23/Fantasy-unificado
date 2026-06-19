from pathlib import Path
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from data_loader import load_workbook_data, SEASONS
from transforms import (
    SEASON_LABELS,
    get_team_options,
    get_visible_seasons,
    build_roster_view,
    format_roster_for_display,
    calculate_main_totals,
    calculate_dev_totals,
)

st.set_page_config(page_title="Elencos Fantasy NBA", layout="wide")
st.title("Elencos Fantasy NBA")
st.caption("Transactions com domínio por time, seleção de players dos dois rosters e save robusto.")

DEFAULT_FILE = Path("roster.xlsx")
TX_SHEET = "transactions"
TX_ITEMS_SHEET = "transaction_times"

@st.cache_data
def cached_load(file_path: str):
    return load_workbook_data(file_path)


def currency(v: float) -> str:
    if pd.isna(v):
        return "-"
    return f"US$ {v:,.2f}"


def display_table(df: pd.DataFrame):
    st.dataframe(df, use_container_width=True, hide_index=True)


def ensure_headers(ws, headers):
    existing = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    if not existing or all(v is None for v in existing):
        ws.append(headers)
        return
    if list(existing) != list(headers):
        raise ValueError(f"Cabeçalhos divergentes em {ws.title}")


def get_next_id(df: pd.DataFrame, col: str, start: int = 1) -> int:
    if df.empty or col not in df.columns:
        return start
    s = pd.to_numeric(df[col], errors="coerce").dropna()
    return int(s.max()) + 1 if not s.empty else start


def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    out = out.where(pd.notna(out), None)
    return out


def load_sheet_df(wb, sheet_name: str) -> pd.DataFrame:
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb[sheet_name]
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows[1:], columns=rows[0])


def save_sheet_df(wb, sheet_name: str, df: pd.DataFrame):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(sheet_name)
    if df.empty:
        return
    df2 = normalize_df(df)
    df2 = df2.astype(object).where(pd.notna(df2), None)
    ws.append(list(df2.columns))
    for _, row in df2.iterrows():
        ws.append([None if pd.isna(v) else v for v in row.tolist()])


def roster_domain_ids(data, team_id: int) -> set:
    ids = set()
    for sheet in ["roster", "development"]:
        df = data.get(sheet, pd.DataFrame())
        if not df.empty and {"team_id", "player_id"}.issubset(df.columns):
            ids |= set(df.loc[df["team_id"].astype(int) == team_id, "player_id"].dropna().astype(int).tolist())
    return ids


def pick_domain_ids(data, team_id: int) -> set:
    picks = data.get("picks", pd.DataFrame())
    if picks.empty:
        return set()
    owner_cols = [c for c in picks.columns if "owner" in c.lower() or "team" in c.lower()]
    if not owner_cols:
        return set()
    owner_col = owner_cols[0]
    id_col = picks.columns[0]
    return set(picks.loc[picks[owner_col].astype(int) == team_id, id_col].astype(str).tolist())


def validate_items(data, from_team_id: int, item_rows: list[dict]) -> list[str]:
    errors = []
    player_ids = roster_domain_ids(data, from_team_id)
    pick_ids = pick_domain_ids(data, from_team_id)
    for i, item in enumerate(item_rows, start=1):
        if item["item_type"] == "player" and item["asset_id"] not in player_ids:
            errors.append(f"Item {i}: jogador fora do domínio do time origem.")
        if item["item_type"] == "pick" and str(item["asset_id"]) not in pick_ids:
            errors.append(f"Item {i}: pick fora do domínio do time origem.")
    return errors


def append_transaction(file_path: str, tx_row: dict, item_rows: list[dict]):
    wb = load_workbook(file_path)
    tx_df = load_sheet_df(wb, TX_SHEET)
    items_df = load_sheet_df(wb, TX_ITEMS_SHEET)
    if tx_df.empty and TX_SHEET not in wb.sheetnames:
        tx_df = pd.DataFrame(columns=list(tx_row.keys()))
    if items_df.empty and TX_ITEMS_SHEET not in wb.sheetnames and item_rows:
        items_df = pd.DataFrame(columns=list(item_rows[0].keys()))
    tx_df = pd.concat([tx_df, pd.DataFrame([tx_row])], ignore_index=True)
    if item_rows:
        items_df = pd.concat([items_df, pd.DataFrame(item_rows)], ignore_index=True)
    save_sheet_df(wb, TX_SHEET, tx_df)
    if item_rows:
        save_sheet_df(wb, TX_ITEMS_SHEET, items_df)
    wb.save(file_path)


def update_rosters(file_path: str, tx_row: dict, item_rows: list[dict]):
    wb = load_workbook(file_path)
    roster_df = load_sheet_df(wb, "roster")
    dev_df = load_sheet_df(wb, "development")
    picks_df = load_sheet_df(wb, "picks")
    from_team_id = int(tx_row["from_team_id"])
    to_team_id = int(tx_row["to_team_id"])
    for item in item_rows:
        if item["item_type"] == "player":
            pid = int(item["asset_id"])
            if item["from_roster_type"] == "MAIN" and not roster_df.empty:
                mask = (roster_df["team_id"].astype(int) == from_team_id) & (roster_df["player_id"].astype(int) == pid)
                roster_df.loc[mask, "team_id"] = to_team_id
            elif item["from_roster_type"] == "DEV" and not dev_df.empty:
                mask = (dev_df["team_id"].astype(int) == from_team_id) & (dev_df["player_id"].astype(int) == pid)
                dev_df.loc[mask, "team_id"] = to_team_id
        elif item["item_type"] == "pick" and not picks_df.empty:
            id_col = picks_df.columns[0]
            owner_cols = [c for c in picks_df.columns if "owner" in c.lower() or "team" in c.lower()]
            if owner_cols:
                owner_col = owner_cols[0]
                mask = picks_df[id_col].astype(str).eq(str(item["asset_id"])) & picks_df[owner_col].astype(int).eq(from_team_id)
                picks_df.loc[mask, owner_col] = to_team_id
    save_sheet_df(wb, "roster", roster_df)
    save_sheet_df(wb, "development", dev_df)
    save_sheet_df(wb, "picks", picks_df)
    wb.save(file_path)


def build_red_flags(df: pd.DataFrame, visible_seasons: list[str]) -> pd.DataFrame:
    out = df.copy()
    for season in visible_seasons:
        sal_col = SEASON_LABELS[season]
        opt_col = f"TO_{season}"
        if sal_col not in out.columns or opt_col not in out.columns:
            continue
        out[f"{sal_col}__red"] = out[opt_col].astype(str).str.strip().str.lower().eq("sim")
        out = out.drop(columns=[opt_col])
    return out


def render_red_table(df: pd.DataFrame, visible_seasons: list[str]) -> str:
    out = df.copy()
    for season in visible_seasons:
        sal_col = SEASON_LABELS[season]
        flag_col = f"{sal_col}__red"
        if sal_col in out.columns and flag_col in out.columns:
            out[sal_col] = out[sal_col].apply(lambda x: f"{x}" if pd.notna(x) else "")
            out = out.drop(columns=[flag_col])
    return out.to_html(escape=False, index=False)

if not DEFAULT_FILE.exists():
    st.error("Arquivo roster.xlsx não encontrado na pasta do projeto.")
    st.stop()

data = cached_load(str(DEFAULT_FILE))
teams = get_team_options(data["teams"])
team_map = data["teams"][["team_id", "team_name"]].drop_duplicates() if not data["teams"].empty else pd.DataFrame(columns=["team_id", "team_name"])
team_lookup = dict(zip(team_map["team_id"], team_map["team_name"])) if not team_map.empty else {}
player_lookup = dict(zip(data["players"]["player_id"], data["players"]["player_name"])) if not data["players"].empty else {}
player_main = data["roster"][["player_id"]].drop_duplicates() if not data["roster"].empty else pd.DataFrame(columns=["player_id"])
player_dev = data["development"][["player_id"]].drop_duplicates() if not data["development"].empty else pd.DataFrame(columns=["player_id"])
player_choices = sorted(set(player_main.get("player_id", pd.Series(dtype=int)).dropna().astype(int).tolist()) | set(player_dev.get("player_id", pd.Series(dtype=int)).dropna().astype(int).tolist()))
player_labels = {pid: f"{pid} - {player_lookup.get(pid, pid)}" for pid in player_choices}
picks_df = data.get("picks", pd.DataFrame())
pick_id_col = picks_df.columns[0] if not picks_df.empty else None
pick_choices = picks_df[pick_id_col].astype(str).tolist() if pick_id_col else []

c1, c2 = st.columns([2, 1])
with c1:
    selected_team_name = st.selectbox("Selecione o time", teams["team_name"].tolist())
with c2:
    selected_start_season = st.selectbox("Temporada inicial", SEASONS, format_func=lambda x: SEASON_LABELS[x])

selected_team_id = int(teams.loc[teams["team_name"] == selected_team_name, "team_id"].iloc[0])
visible_seasons = get_visible_seasons(selected_start_season)


main_team_df = data["roster"].loc[data["roster"]["team_id"] == selected_team_id].copy()
dev_team_df = data["development"].loc[data["development"]["team_id"] == selected_team_id].copy()
main_roster_raw = build_roster_view(main_team_df, data["players"], selected_team_id, "MAIN", visible_seasons)
main_roster = format_roster_for_display(main_roster_raw, visible_seasons)
main_totals = calculate_main_totals(main_team_df, data["fines"], selected_team_id, visible_seasons)
dev_roster_raw = build_roster_view(dev_team_df, data["players"], selected_team_id, "DEV", visible_seasons)
dev_roster = format_roster_for_display(dev_roster_raw, visible_seasons)
dev_totals = calculate_dev_totals(dev_team_df, visible_seasons)



st.subheader("Elenco principal")
display_main = build_red_flags(main_roster, visible_seasons)
for season in visible_seasons:
    label = SEASON_LABELS[season]
    if label in display_main.columns:
        display_main[label] = display_main[label].apply(currency)
display_table(display_main)

st.subheader("Totalizadores do elenco principal")
main_totals_display = main_totals.copy()
for col in ["Salários", "Multas", "Cap restante"]:
    if col in main_totals_display.columns:
        main_totals_display[col] = main_totals_display[col].apply(currency)
st.dataframe(main_totals_display, use_container_width=True, hide_index=True)

st.subheader("Liga de desenvolvimento")
display_dev = build_red_flags(dev_roster, visible_seasons)
for season in visible_seasons:
    label = SEASON_LABELS[season]
    if label in display_dev.columns:
        display_dev[label] = display_dev[label].apply(currency)
display_table(display_dev)

st.subheader("Totalizadores da development")
dev_totals_display = dev_totals.copy()
for col in ["Salários", "Cap restante"]:
    if col in dev_totals_display.columns:
        dev_totals_display[col] = dev_totals_display[col].apply(currency)
st.dataframe(dev_totals_display, use_container_width=True, hide_index=True)
