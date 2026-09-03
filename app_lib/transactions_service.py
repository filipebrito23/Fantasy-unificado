import pandas as pd
from openpyxl import load_workbook
import streamlit as st

from app_lib.excel_utils import load_sheet_df, save_sheet_df, ensure_unique_columns


TX_SHEET = "transactions"
TX_ITEMS_SHEET = "transactionitems"



def _to_int(value):
    try:
        return int(value) if value is not None and str(value).strip() != "" else None
    except Exception:
        return None



def _normalize_roster_type(value):
    return str(value or "").strip().upper()



def _item_value(item, *keys, default=None):
    for key in keys:
        if isinstance(item, dict) and key in item:
            val = item.get(key)
            if val is not None and str(val).strip() != "":
                return val
    return default



def roster_domain_ids(data, team_id: int, roster_type: str | None = None) -> set:
    ids = set()
    sheet_names = ["roster", "development"]
    roster_type = _normalize_roster_type(roster_type)
    if roster_type == "MAIN":
        sheet_names = ["roster"]
    elif roster_type == "DEV":
        sheet_names = ["development"]


    for sheet in sheet_names:
        df = data.get(sheet, pd.DataFrame())
        if df.empty or not {"team_id", "player_id"}.issubset(df.columns):
            continue
        team_ids = pd.to_numeric(df["team_id"], errors="coerce")
        player_ids = pd.to_numeric(df["player_id"], errors="coerce")
        mask = team_ids.eq(team_id) & player_ids.notna()
        ids |= set(player_ids.loc[mask].astype(int).tolist())
    return ids



def pick_domain_ids(data, team_id: int) -> set:
    """
    Retorna as picks que pertencem ATUALMENTE ao time.

    A regra oficial é sempre baseada em current_team_owner_id.
    Não usa original_team_pick_id, pois a pick pode ter sido adquirida
    em uma troca e, mesmo assim, deve aparecer para o dono atual.
    """
    picks = data.get("picks", pd.DataFrame())

    if picks.empty:
        return set()

    required_columns = {"pick_id", "current_team_owner_id"}

    if not required_columns.issubset(picks.columns):
        return set()

    owner_ids = pd.to_numeric(
        picks["current_team_owner_id"],
        errors="coerce",
    )

    valid_mask = owner_ids.notna() & owner_ids.eq(int(team_id))

    return set(
        picks.loc[valid_mask, "pick_id"]
        .dropna()
        .astype(str)
        .str.strip()
        .tolist()
    )



def validate_items(data, from_team_id: int, item_rows: list[dict]) -> list[str]:
    errors = []
    for i, item in enumerate(item_rows, start=1):
        item_type = str(_item_value(item, "item_type", "itemtype", default="")).strip().lower()
        asset_id = _item_value(item, "asset_id", "assetid")
        from_roster_type = _item_value(item, "from_roster_type", "fromrostertype")
        player_ids = roster_domain_ids(data, from_team_id, from_roster_type)
        pick_ids = pick_domain_ids(data, from_team_id)
        if item_type == "player":
            asset_id_int = _to_int(asset_id)
            if asset_id_int is None or asset_id_int not in player_ids:
                errors.append(f"Item {i}: jogador fora do domínio do time origem.")
        elif item_type == "pick":
            if str(asset_id) not in pick_ids:
                errors.append(f"Item {i}: pick fora do domínio do time origem.")
        else:
            errors.append(f"Item {i}: tipo de asset inválido.")
    return errors



def validate_items_bilateral(data, item_rows: list[dict], transaction_type: str | None = None) -> list[str]:
    errors = []
    tx_type = str(transaction_type or "").strip().upper()
    for i, item in enumerate(item_rows, start=1):
        item_type = str(_item_value(item, "item_type", "itemtype", default="")).strip().lower()
        asset_id = _item_value(item, "asset_id", "assetid")
        from_team_id = _to_int(_item_value(item, "from_team_id", "fromteamid"))
        to_team_id = _to_int(_item_value(item, "to_team_id", "toteamid"))
        from_roster_type = _normalize_roster_type(_item_value(item, "from_roster_type", "fromrostertype"))
        to_roster_type = _normalize_roster_type(_item_value(item, "to_roster_type", "torostertype"))


        if tx_type == "TRADE":
            if from_team_id is None or to_team_id is None:
                errors.append(f"Item {i}: trade exige time de origem e destino.")
                continue
            if from_team_id == to_team_id:
                errors.append(f"Item {i}: trade exige times diferentes.")
                continue
        elif tx_type in {"WAIVE", "DISPENSA", "DISMISS", "DROP"}:
            if from_team_id is None:
                errors.append(f"Item {i}: dispensa exige time de origem.")
                continue
        elif tx_type in {"ADD", "SIGN", "ASSINATURA"}:
            if to_team_id is None:
                errors.append(f"Item {i}: adição exige time de destino.")
                continue
        elif tx_type in {"MOVE", "CALLUP", "SENDDOWN", "PROMOTION"}:
            if from_team_id is None:
                errors.append(f"Item {i}: movimentação interna exige origem e destino.")
                continue
            if to_team_id is None:
                to_team_id = from_team_id


        same_team = from_team_id is not None and to_team_id is not None and from_team_id == to_team_id


        if item_type == "player":
            asset_id_int = _to_int(asset_id)
            if asset_id_int is None:
                errors.append(f"Item {i}: jogador inválido.")
                continue
            if from_team_id is not None:
                player_ids = roster_domain_ids(data, from_team_id, from_roster_type)
                if asset_id_int not in player_ids:
                    errors.append(f"Item {i}: jogador fora do domínio do time origem.")
                    continue
            if tx_type in {"MOVE", "CALLUP", "SENDDOWN", "PROMOTION"} or same_team:
                if from_roster_type not in {"MAIN", "DEV"} or to_roster_type not in {"MAIN", "DEV"}:
                    errors.append(f"Item {i}: movimentação interna exige MAIN/DEV válidos.")
                elif from_roster_type == to_roster_type:
                    errors.append(f"Item {i}: movimentação interna deve trocar entre MAIN e DEV.")
        elif item_type == "pick":
            if from_team_id is None:
                errors.append(f"Item {i}: pick exige time de origem.")
                continue
            pick_ids = pick_domain_ids(data, from_team_id)
            if str(asset_id) not in pick_ids:
                errors.append(f"Item {i}: pick fora do domínio do time origem.")
            if tx_type in {"MOVE", "CALLUP", "SENDDOWN", "PROMOTION"} or same_team:
                errors.append(f"Item {i}: pick não pode ser movimentada dentro do mesmo time.")
            if tx_type in {"WAIVE", "DISPENSA", "DISMISS", "DROP"}:
                errors.append(f"Item {i}: dispensa não se aplica a pick.")
        else:
            errors.append(f"Item {i}: tipo de asset inválido.")
    return errors



def append_transaction(file_path: str, tx_row: dict, item_rows: list[dict]):
    wb = load_workbook(file_path)
    tx_df = ensure_unique_columns(load_sheet_df(wb, TX_SHEET))
    items_df = ensure_unique_columns(load_sheet_df(wb, TX_ITEMS_SHEET))
    if tx_df.empty:
        tx_df = pd.DataFrame(columns=list(tx_row.keys()))
    if items_df.empty and item_rows:
        items_df = pd.DataFrame(columns=list(item_rows[0].keys()))
    tx_df = ensure_unique_columns(tx_df)
    items_df = ensure_unique_columns(items_df)


    tx_id_col = next((c for c in tx_df.columns if str(c).strip().lower() == "transaction_id"), None)
    item_tx_id_col = next((c for c in items_df.columns if str(c).strip().lower() == "transaction_id"), None)
    if tx_id_col is None:
        tx_id_col = "transaction_id"
        if tx_id_col not in tx_df.columns:
            tx_df[tx_id_col] = pd.Series(dtype="Int64")
    if item_tx_id_col is None:
        item_tx_id_col = "transaction_id"
        if item_tx_id_col not in items_df.columns:
            items_df[item_tx_id_col] = pd.Series(dtype="Int64")


    existing_ids = pd.to_numeric(tx_df[tx_id_col], errors="coerce").dropna()
    next_tx_id = int(existing_ids.max()) + 1 if not existing_ids.empty else 1


    base_tx = dict(tx_row)
    base_tx[tx_id_col] = next_tx_id
    new_tx_rows = [base_tx]
    new_item_rows = []
    for item in item_rows or []:
        row = dict(item)
        row[item_tx_id_col] = next_tx_id
        new_item_rows.append(row)


    tx_df = pd.concat([tx_df, ensure_unique_columns(pd.DataFrame(new_tx_rows))], ignore_index=True)
    if new_item_rows:
        items_df = pd.concat([items_df, ensure_unique_columns(pd.DataFrame(new_item_rows))], ignore_index=True)


    save_sheet_df(wb, TX_SHEET, tx_df)
    if not items_df.empty:
        save_sheet_df(wb, TX_ITEMS_SHEET, items_df)
    wb.save(file_path)



def update_rosters(file_path: str, tx_row: dict, item_rows: list[dict]):
    wb = load_workbook(file_path)

    roster_df = load_sheet_df(wb, "roster")
    dev_df = load_sheet_df(wb, "development")
    picks_df = load_sheet_df(wb, "picks")

    tx_type = str(
        tx_row.get(
            "transaction_type",
            tx_row.get("transactiontype", ""),
        )
    ).strip().upper()

    default_from_team_id = _to_int(
        tx_row.get(
            "from_team_id",
            tx_row.get("fromteamid"),
        )
    )

    default_to_team_id = _to_int(
        tx_row.get(
            "to_team_id",
            tx_row.get("toteamid"),
        )
    )

    def num(series):
        return pd.to_numeric(series, errors="coerce")

    def pick_id_key(value):
        """
        Normaliza ID de pick para comparação.

        Exemplos tratados como equivalentes:
        P2026R1T05
        P2026_R1_T5
        p2026-r1-t05
        """
        if value is None or pd.isna(value):
            return ""

        return (
            str(value)
            .strip()
            .upper()
            .replace(" ", "")
            .replace("_", "")
            .replace("-", "")
        )

    for item in item_rows:
        item_type = str(
            _item_value(
                item,
                "item_type",
                "itemtype",
                default="",
            )
        ).strip().lower()

        asset_id = _item_value(item, "asset_id", "assetid")

        item_from_team = _to_int(
            _item_value(
                item,
                "from_team_id",
                "fromteamid",
                default=default_from_team_id,
            )
        )

        item_to_team = _to_int(
            _item_value(
                item,
                "to_team_id",
                "toteamid",
                default=default_to_team_id,
            )
        )

        # ---------------------------------------------------------
        # PICKS
        # ---------------------------------------------------------
        if item_type == "pick":
            # Picks não se aplicam a dispensa.
            if tx_type in {"WAIVE", "DISPENSA", "DISMISS", "DROP"}:
                continue

            # Picks não podem ser movimentadas entre MAIN e DEV.
            if tx_type in {
                "MOVE",
                "CALLUP",
                "SENDDOWN",
                "PROMOTION",
            }:
                continue

            # Para uma trade, a pick precisa sair de um time e ir para outro.
            if item_from_team is None or item_to_team is None:
                print(
                    "PICK UPDATE DEBUG: origem ou destino inválido. "
                    f"pick={asset_id}; from={item_from_team}; "
                    f"to={item_to_team}"
                )
                continue

            if picks_df.empty:
                print(
                    "PICK UPDATE DEBUG: aba picks está vazia. "
                    f"pick={asset_id}"
                )
                continue

            required_pick_columns = {
                "pick_id",
                "current_team_owner_id",
            }

            if not required_pick_columns.issubset(picks_df.columns):
                print(
                    "PICK UPDATE DEBUG: colunas obrigatórias "
                    "não encontradas na aba picks. "
                    f"Colunas atuais: {list(picks_df.columns)}"
                )
                continue

            wanted_pick_key = pick_id_key(asset_id)

            if not wanted_pick_key:
                print("PICK UPDATE DEBUG: pick_id inválido.")
                continue

            current_owner_ids = pd.to_numeric(
                picks_df["current_team_owner_id"],
                errors="coerce",
            )

            pick_keys = picks_df["pick_id"].apply(pick_id_key)

            mask = (
                pick_keys.eq(wanted_pick_key)
                & current_owner_ids.eq(item_from_team)
            )

            if not mask.any():
                matching_pick = pick_keys.eq(wanted_pick_key)

                if matching_pick.any():
                    owners_found = (
                        picks_df.loc[
                            matching_pick,
                            "current_team_owner_id",
                        ]
                        .dropna()
                        .astype(str)
                        .tolist()
                    )

                    print(
                        "PICK UPDATE DEBUG: a pick existe, mas "
                        "não pertence ao time de origem informado. "
                        f"pick={asset_id}; "
                        f"origem_esperada={item_from_team}; "
                        f"donos_encontrados={owners_found}"
                    )
                else:
                    print(
                        "PICK UPDATE DEBUG: pick não encontrada. "
                        f"pick={asset_id}; "
                        f"chave_normalizada={wanted_pick_key}"
                    )

                continue

            picks_df.loc[
                mask,
                "current_team_owner_id",
            ] = int(item_to_team)

            print(
                "PICK UPDATE DEBUG: propriedade atualizada. "
                f"pick={asset_id}; "
                f"{item_from_team} -> {item_to_team}"
            )

            # Não deixa a lógica de jogador processar uma pick.
            continue

        # ---------------------------------------------------------
        # JOGADORES
        # ---------------------------------------------------------
        if item_type != "player":
            continue

        pid = _to_int(asset_id)

        if pid is None:
            print(
                "PLAYER UPDATE DEBUG: player_id inválido. "
                f"asset={asset_id}"
            )
            continue

        from_rt = _normalize_roster_type(
            _item_value(
                item,
                "from_roster_type",
                "fromrostertype",
            )
        )

        to_rt = _normalize_roster_type(
            _item_value(
                item,
                "to_roster_type",
                "torostertype",
            )
        )

        # ---------------------------------------------------------
        # TRADE: jogador sai do MAIN da origem e entra no destino.
        # Caso to_roster_type esteja vazio, o destino padrão é MAIN.
        # ---------------------------------------------------------
        if tx_type == "TRADE":
            if item_from_team is None or item_to_team is None:
                continue

            if to_rt not in {"MAIN", "DEV"}:
                to_rt = "MAIN"

            source_df = roster_df if from_rt != "DEV" else dev_df

            if source_df.empty or not {
                "team_id",
                "player_id",
            }.issubset(source_df.columns):
                print(
                    "TRADE REMOVE DEBUG: fonte de jogador inválida. "
                    f"player={pid}; origem={item_from_team}; "
                    f"roster={from_rt}"
                )
                continue

            source_mask = (
                num(source_df["team_id"]).eq(item_from_team)
                & num(source_df["player_id"]).eq(pid)
            )

            if not source_mask.any():
                print(
                    "TRADE REMOVE DEBUG: não encontrou "
                    f"player {pid} no {from_rt or 'MAIN'} "
                    f"do time {item_from_team}"
                )
                continue

            moved = source_df.loc[source_mask].copy()
            source_df = source_df.loc[~source_mask].copy()

            moved.loc[:, "team_id"] = item_to_team

            if from_rt == "DEV":
                dev_df = source_df
            else:
                roster_df = source_df

            if to_rt == "DEV":
                dev_df = pd.concat(
                    [dev_df, moved],
                    ignore_index=True,
                )
            else:
                roster_df = pd.concat(
                    [roster_df, moved],
                    ignore_index=True,
                )

            continue

        # ---------------------------------------------------------
        # WAIVE: remove do roster correspondente.
        # ---------------------------------------------------------
        if tx_type in {"WAIVE", "DISPENSA", "DISMISS", "DROP"}:
            if item_from_team is None:
                continue

            if from_rt == "DEV" and not dev_df.empty:
                mask = (
                    num(dev_df["team_id"]).eq(item_from_team)
                    & num(dev_df["player_id"]).eq(pid)
                )
                dev_df = dev_df.loc[~mask].copy()
            elif not roster_df.empty:
                mask = (
                    num(roster_df["team_id"]).eq(item_from_team)
                    & num(roster_df["player_id"]).eq(pid)
                )
                roster_df = roster_df.loc[~mask].copy()

            continue

        # ---------------------------------------------------------
        # ADD / SIGN: mantém a lógica existente de mover jogador.
        # ---------------------------------------------------------
        if tx_type in {"ADD", "SIGN", "ASSINATURA"}:
            if item_from_team is None or item_to_team is None:
                continue

            source_df = dev_df if from_rt == "DEV" else roster_df
            target_df = dev_df if to_rt == "DEV" else roster_df

            if source_df.empty or not {
                "team_id",
                "player_id",
            }.issubset(source_df.columns):
                continue

            source_mask = (
                num(source_df["team_id"]).eq(item_from_team)
                & num(source_df["player_id"]).eq(pid)
            )

            if not source_mask.any():
                continue

            moved = source_df.loc[source_mask].copy()
            source_df = source_df.loc[~source_mask].copy()

            moved.loc[:, "team_id"] = item_to_team
            target_df = pd.concat(
                [target_df, moved],
                ignore_index=True,
            )

            if from_rt == "DEV":
                dev_df = source_df
            else:
                roster_df = source_df

            if to_rt == "DEV":
                dev_df = target_df
            else:
                roster_df = target_df

            continue

        # ---------------------------------------------------------
        # MOVE / CALLUP / SENDDOWN / PROMOTION:
        # movimenta jogador entre MAIN e DEV dentro do mesmo time.
        # ---------------------------------------------------------
        if tx_type in {
            "MOVE",
            "CALLUP",
            "SENDDOWN",
            "PROMOTION",
        }:
            if (
                from_rt not in {"MAIN", "DEV"}
                or to_rt not in {"MAIN", "DEV"}
                or item_from_team is None
                or item_to_team is None
            ):
                continue

            source_df = dev_df if from_rt == "DEV" else roster_df
            target_df = dev_df if to_rt == "DEV" else roster_df

            if source_df.empty or not {
                "team_id",
                "player_id",
            }.issubset(source_df.columns):
                continue

            source_mask = (
                num(source_df["team_id"]).eq(item_from_team)
                & num(source_df["player_id"]).eq(pid)
            )

            if not source_mask.any():
                continue

            moved = source_df.loc[source_mask].copy()
            source_df = source_df.loc[~source_mask].copy()

            moved.loc[:, "team_id"] = item_to_team

            target_df = pd.concat(
                [target_df, moved],
                ignore_index=True,
            )

            if from_rt == "DEV":
                dev_df = source_df
            else:
                roster_df = source_df

            if to_rt == "DEV":
                dev_df = target_df
            else:
                roster_df = target_df

    save_sheet_df(wb, "roster", roster_df)
    save_sheet_df(wb, "development", dev_df)
    save_sheet_df(wb, "picks", picks_df)

    wb.save(file_path)


def compact_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).strip().lower().replace("_", "").replace(" ", "") for c in out.columns]
    out = out.loc[:, ~out.columns.duplicated()].copy()
    return out



def build_transactions_history(
    tx_df: pd.DataFrame,
    items_df: pd.DataFrame,
    selected_team_id: int,
    team_lookup: dict,
    player_lookup: dict,
) -> pd.DataFrame:
    """
    Monta o histórico de transactions de um time, ligando cada
    transaction aos seus itens por transaction_id.

    A função normaliza IDs como 58, 58.0 e "58" para o mesmo formato,
    evitando que itens fiquem desconectados da transaction.
    """
    if tx_df is None or tx_df.empty:
        return pd.DataFrame()

    def normalize_transaction_id(value) -> str | None:
        if value is None or pd.isna(value):
            return None

        try:
            return str(int(float(value)))
        except (TypeError, ValueError):
            text_value = str(value).strip()
            return text_value if text_value else None

    def normalize_team_id(value) -> int | None:
        if value is None or pd.isna(value):
            return None

        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None

    tx = compact_columns(tx_df)
    items = (
        compact_columns(items_df)
        if items_df is not None and not items_df.empty
        else pd.DataFrame()
    )

    tx = tx.loc[:, ~tx.columns.duplicated()].copy()

    required_tx_columns = {
        "transactionid",
        "fromteamid",
        "toteamid",
    }

    if not required_tx_columns.issubset(tx.columns):
        return pd.DataFrame()

    selected_team_id = int(selected_team_id)

    tx["__transaction_key__"] = tx["transactionid"].apply(
        normalize_transaction_id
    )
    tx["__from_team_id__"] = tx["fromteamid"].apply(
        normalize_team_id
    )
    tx["__to_team_id__"] = tx["toteamid"].apply(
        normalize_team_id
    )

    tx = tx.loc[
        tx["__from_team_id__"].eq(selected_team_id)
        | tx["__to_team_id__"].eq(selected_team_id)
    ].copy()

    if tx.empty:
        return pd.DataFrame()

    tx["from_team"] = tx["__from_team_id__"].map(team_lookup).fillna(
        tx["fromteamid"]
    )
    tx["to_team"] = tx["__to_team_id__"].map(team_lookup).fillna(
        tx["toteamid"]
    )

    items_by_transaction: dict[str, list[dict]] = {}

    if not items.empty and "transactionid" in items.columns:
        items = items.copy()

        items["__transaction_key__"] = items["transactionid"].apply(
            normalize_transaction_id
        )
        items["__from_team_id__"] = items["fromteamid"].apply(
            normalize_team_id
        )
        items["__to_team_id__"] = items["toteamid"].apply(
            normalize_team_id
        )

        valid_transaction_keys = set(
            tx["__transaction_key__"].dropna().tolist()
        )

        items = items.loc[
            items["__transaction_key__"].isin(valid_transaction_keys)
        ].copy()

        for _, item in items.iterrows():
            transaction_key = item.get("__transaction_key__")

            if not transaction_key:
                continue

            item_type = str(
                item.get("itemtype", "")
            ).strip().lower()

            asset_id = item.get("assetid")

            from_team_id = item.get("__from_team_id__")
            to_team_id = item.get("__to_team_id__")

            from_roster_type = str(
                item.get("fromrostertype", "") or ""
            ).strip().upper()

            to_roster_type = str(
                item.get("torostertype", "") or ""
            ).strip().upper()

            if item_type == "player":
                player_id = normalize_team_id(asset_id)

                if player_id is None:
                    asset_label = str(asset_id)
                else:
                    asset_label = player_lookup.get(
                        player_id,
                        f"Jogador #{player_id}",
                    )

            elif item_type == "pick":
                asset_label = str(asset_id).strip()

            else:
                asset_label = str(asset_id).strip()

            roster_label = ""

            if from_roster_type or to_roster_type:
                roster_label = (
                    f" [{from_roster_type or '-'}"
                    f" → {to_roster_type or '-'}]"
                )

            items_by_transaction.setdefault(
                str(transaction_key),
                [],
            ).append(
                {
                    "from_team_id": from_team_id,
                    "to_team_id": to_team_id,
                    "label": f"{asset_label}{roster_label}",
                }
            )

    sent_values: list[str] = []
    received_values: list[str] = []
    all_items_values: list[str] = []

    for _, tx_row in tx.iterrows():
        transaction_key = tx_row.get("__transaction_key__")
        transaction_items = items_by_transaction.get(
            str(transaction_key),
            [],
        )

        sent_items = [
            item["label"]
            for item in transaction_items
            if item.get("from_team_id") == selected_team_id
        ]

        received_items = [
            item["label"]
            for item in transaction_items
            if item.get("to_team_id") == selected_team_id
        ]

        all_items = [
            item["label"]
            for item in transaction_items
        ]

        sent_values.append(
            " | ".join(sent_items) if sent_items else "-"
        )
        received_values.append(
            " | ".join(received_items) if received_items else "-"
        )
        all_items_values.append(
            " | ".join(all_items) if all_items else "-"
        )

    tx["enviado"] = sent_values
    tx["recebido"] = received_values
    tx["itens"] = all_items_values

    tx = tx.rename(
        columns={
            "transactionid": "transaction_id",
            "transactiontype": "transaction_type",
            "transactiondate": "transaction_date",
            "initiatedby": "initiated_by",
        }
    )

    tx = tx.loc[:, ~tx.columns.duplicated()].copy()

    preferred_columns = [
        "transaction_date",
        "season",
        "from_team",
        "to_team",
        "enviado",
        "recebido",
        "notes",
    ]

    existing_columns = [
        column
        for column in preferred_columns
        if column in tx.columns
    ]

    sort_columns = [
        column
        for column in ["transaction_date", "transaction_id"]
        if column in tx.columns
    ]

    if sort_columns:
        tx = tx.sort_values(
            by=sort_columns,
            ascending=False,
            kind="stable",
        )

    return tx[existing_columns].reset_index(drop=True)