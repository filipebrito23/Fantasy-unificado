# transactions_service.py

import pandas as pd
from openpyxl import load_workbook

from excel_utils import load_sheet_df, save_sheet_df

TX_SHEET = "transactions"
TX_ITEMS_SHEET = "transactionitems"


def roster_domain_ids(data, team_id: int) -> set:
    ids = set()
    for sheet in ["roster", "development"]:
        df = data.get(sheet, pd.DataFrame())
        if not df.empty and {"team_id", "player_id"}.issubset(df.columns):
            ids |= set(
                df.loc[df["team_id"].astype(int) == team_id, "player_id"]
                .dropna()
                .astype(int)
                .tolist()
            )
    return ids


def pick_domain_ids(data, team_id: int) -> set:
    picks = data.get("picks", pd.DataFrame())
    if picks.empty:
        return set()

    owner_cols = [c for c in picks.columns if "owner" in c.lower() or "team" in c.lower()]
    if not owner_cols:
        return set()

    owner_col = owner_cols[0]
    id_col = picks.columns[0]

    return set(
        picks.loc[picks[owner_col].astype(int) == team_id, id_col]
        .astype(str)
        .tolist()
    )


def validate_items(data, from_team_id: int, item_rows: list[dict]) -> list[str]:
    errors = []
    player_ids = roster_domain_ids(data, from_team_id)
    pick_ids = pick_domain_ids(data, from_team_id)

    for i, item in enumerate(item_rows, start=1):
        if item["item_type"] == "player" and item["asset_id"] not in player_ids:
            errors.append(f"Item {i}: jogador fora do domínio do time origem.")
        if item["item_type"] == "pick" and str(item["asset_id"]) not in pick_ids:
            errors.append(f"Item {i}: pick fora do domínio do time origem.")

    return errors


def append_transaction(file_path: str, tx_row: dict, item_rows: list[dict]):
    wb = load_workbook(file_path)

    tx_df = load_sheet_df(wb, TX_SHEET)
    items_df = load_sheet_df(wb, TX_ITEMS_SHEET)

    if tx_df.empty and TX_SHEET not in wb.sheetnames:
        tx_df = pd.DataFrame(columns=list(tx_row.keys()))

    if items_df.empty and TX_ITEMS_SHEET not in wb.sheetnames and item_rows:
        items_df = pd.DataFrame(columns=list(item_rows[0].keys()))

    tx_df = pd.concat([tx_df, pd.DataFrame([tx_row])], ignore_index=True)

    if item_rows:
        items_df = pd.concat([items_df, pd.DataFrame(item_rows)], ignore_index=True)

    save_sheet_df(wb, TX_SHEET, tx_df)
    if item_rows:
        save_sheet_df(wb, TX_ITEMS_SHEET, items_df)

    wb.save(file_path)


def update_rosters(file_path: str, tx_row: dict, item_rows: list[dict]):
    wb = load_workbook(file_path)

    roster_df = load_sheet_df(wb, "roster")
    dev_df = load_sheet_df(wb, "development")
    picks_df = load_sheet_df(wb, "picks")

    from_team_id = tx_row.get("from_team_id", tx_row.get("fromteamid"))
    to_team_id = tx_row.get("to_team_id", tx_row.get("toteamid"))

    if from_team_id is None or to_team_id is None:
        raise ValueError("Transaction sem from_team_id/to_team_id válidos.")

    from_team_id = int(from_team_id)
    to_team_id = int(to_team_id)

    for item in item_rows:
        item_type = item.get("item_type", item.get("itemtype"))
        asset_id = item.get("asset_id", item.get("assetid"))
        from_roster_type = item.get("from_roster_type", item.get("fromrostertype"))
        to_roster_type = item.get("to_roster_type", item.get("torostertype"))

        if item_type == "player":
            if asset_id is None:
                continue

            pid = int(asset_id)

            if str(from_roster_type).strip().upper() == "MAIN" and not roster_df.empty:
                mask = (
                    (pd.to_numeric(roster_df["team_id"], errors="coerce") == from_team_id)
                    & (pd.to_numeric(roster_df["player_id"], errors="coerce") == pid)
                )
                roster_df.loc[mask, "team_id"] = to_team_id

            elif str(from_roster_type).strip().upper() == "DEV" and not dev_df.empty:
                mask = (
                    (pd.to_numeric(dev_df["team_id"], errors="coerce") == from_team_id)
                    & (pd.to_numeric(dev_df["player_id"], errors="coerce") == pid)
                )
                dev_df.loc[mask, "team_id"] = to_team_id

        elif item_type == "pick" and not picks_df.empty:
            if asset_id is None:
                continue

            id_col = picks_df.columns[0]
            owner_cols = [c for c in picks_df.columns if "owner" in c.lower() or "team" in c.lower()]

            if owner_cols:
                owner_col = owner_cols[0]
                mask = (
                    picks_df[id_col].astype(str).eq(str(asset_id))
                    & pd.to_numeric(picks_df[owner_col], errors="coerce").eq(from_team_id)
                )
                picks_df.loc[mask, owner_col] = to_team_id

    save_sheet_df(wb, "roster", roster_df)
    save_sheet_df(wb, "development", dev_df)
    save_sheet_df(wb, "picks", picks_df)

    wb.save(file_path)


def compact_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [
        str(c).strip().lower().replace("_", "").replace(" ", "")
        for c in out.columns
    ]
    return out


def compact_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [
        str(c).strip().lower().replace("_", "").replace(" ", "")
        for c in out.columns
    ]
    return out


def build_transactions_history(
    tx_df: pd.DataFrame,
    items_df: pd.DataFrame,
    selected_team_id: int,
    team_lookup: dict,
    player_lookup: dict,
) -> pd.DataFrame:
    if tx_df is None or tx_df.empty:
        return pd.DataFrame()

    tx = compact_columns(tx_df)

    items = items_df.copy() if items_df is not None else pd.DataFrame()
    if not items.empty:
        items = compact_columns(items)

    if not {"fromteamid", "toteamid"}.issubset(tx.columns):
        return pd.DataFrame()

    team_mask = (
        tx["fromteamid"].astype(str).eq(str(selected_team_id))
        | tx["toteamid"].astype(str).eq(str(selected_team_id))
    )
    tx = tx.loc[team_mask].copy()

    if tx.empty:
        return pd.DataFrame()

    tx["from_team"] = tx["fromteamid"].map(team_lookup).fillna(tx["fromteamid"])
    tx["to_team"] = tx["toteamid"].map(team_lookup).fillna(tx["toteamid"])

    if items.empty or "transactionid" not in items.columns:
        tx["items_summary"] = "-"
    else:
        items = items.loc[
            items["transactionid"].astype(str).isin(tx["transactionid"].astype(str))
        ].copy()

        def describe_item(row):
            item_type = str(row.get("itemtype", "")).upper()
            asset_id = row.get("assetid")

            if item_type == "PLAYER":
                try:
                    pid = int(asset_id)
                except Exception:
                    pid = asset_id
                asset_label = player_lookup.get(pid, asset_id)
            else:
                asset_label = asset_id

            from_rt = row.get("fromrostertype")
            to_rt = row.get("torostertype")

            roster_part = ""
            if pd.notna(from_rt) or pd.notna(to_rt):
                roster_part = f" ({from_rt or '-'} → {to_rt or '-'})"

            return f"{item_type}: {asset_label}{roster_part}"

        if not items.empty:
            items["item_desc"] = items.apply(describe_item, axis=1)
            items_grouped = (
                items.groupby("transactionid")["item_desc"]
                .apply(lambda s: " | ".join(s.astype(str)))
                .reset_index()
            )
            tx = tx.merge(items_grouped, on="transactionid", how="left")
            tx = tx.rename(columns={"item_desc": "items_summary"})
        else:
            tx["items_summary"] = "-"

    tx = tx.rename(columns={
        "transactionid": "transaction_id",
        "transactiontype": "transaction_type",
        "transactiondate": "transaction_date",
        "initiatedby": "initiated_by",
    })

    preferred_cols = [
        "transaction_id",
        "transaction_date",
        "transaction_type",
        "season",
        "from_team",
        "to_team",
        "initiated_by",
        "status",
        "items_summary",
        "notes",
    ]
    existing_cols = [c for c in preferred_cols if c in tx.columns]

    sort_cols = [c for c in ["transaction_date", "transaction_id"] if c in tx.columns]
    if sort_cols:
        tx = tx.sort_values(by=sort_cols, ascending=False)

    return tx[existing_cols]